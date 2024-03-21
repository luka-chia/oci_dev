#coding: utf-8
# OCI Instance Information Collection Tool
# - Description: Try to collect all instances information(including Compute/Autonomouse DB/DBCS/MySQL instances)
# - Version 0.0.2
# - Date 2022.10.27
# - Author: Tam Tan

import oci
import argparse
import sys
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, colors, Alignment

# Deinfe a functino to generate the config dict for a secific profile
def generateClientConfig(user, fingerprint, keyfile, tenancy, region):
    clientConfig = {
        "user": user, 
        "key_file": keyfile,
        "fingerprint": fingerprint,
        "tenancy": tenancy,
        "region": region
    }
    return clientConfig

# Begin to define a entity for compartment
class CompartmentEntity:
    def __init__(self, currentId, parentId, currentName, parentName, level, fullname):
        self.level = level
        self.id = currentId
        self.pid = parentId
        self.name = currentName
        self.pname = parentName
        
        if self.level == 0:
            self.fullname = currentName
        if self.level == 1:
            self.fullname = parentName + " > " + currentName
        elif self.level >= 2:
            self.fullname = fullname + " > " + currentName
        
    def __str__(self):
        return "{'id': '%s', 'pid': '%s', 'name': '%s', 'fullname:': '%s' ,'pname': '%s','level': %s}" % (self.id, self.pid, self.name, self.fullname, self.pname, self.level)
    
# End definition

def isIgnoreCompartment(checkCompartmentList, currentCompartmentEntity):
    ignore = False
    # Start to check on assiged compartment list
    if checkCompartmentList:
        if currentCompartmentEntity.name not in checkCompartmentList:
            ignore = True
            
        for chkName in checkCompartmentList:
            if currentCompartmentEntity.fullname.count(chkName) > 0:
                ignore = False
                break
        # End to check on assiged compartment list
        
    return ignore

# Define a list to store all compartments
totalCompartments = []
# Define a clients config list<dict>
clientConfigList = []

all_resource_list = list()
instance_count_group_by_compartment = dict()
loadLalance_count_group_by_compartment = dict()
mysql_count_group_by_compartment = dict()
waf_count_group_by_compartment = dict()
dbcs_count_group_by_compartment = dict()
storage_count_group_by_compartment = dict()

def initClientContext():
    # Init OCI context
    # Make sure that the default profile exist: ./oci/config [DEFAULT]
    defaultProfile="DEFAULT"
    global totalCompartments
    global instance_count_group_by_compartment
    global loadLalance_count_group_by_compartment
    global mysql_count_group_by_compartment
    global waf_count_group_by_compartment
    global dbcs_count_group_by_compartment
    global storage_count_group_by_compartment

    # Load the DEFAULT profile from file: ./oci/config
    config = oci.config.from_file(profile_name=defaultProfile)

    # List all the DEFAULT profile parameter values
    defaultProfileUser = config["user"]
    defaultProfileFingerprint = config["fingerprint"]
    defaultProfileKeyFile = config["key_file"]
    defaultProfileTenancy = config["tenancy"]
    defaultProfileRegion = config["region"]

    # Init a default client via default profile
    client = oci.identity.IdentityClient(config)

    # Get all subscribed regions
    subscribedRegionList = client.list_region_subscriptions(defaultProfileTenancy).data
    print("The number of subscribed regions: ", len(subscribedRegionList))

    # Get the root compartment - Level 0
    rootCompartment = client.get_compartment(compartment_id=defaultProfileTenancy).data
    rootCompartmentEntity = CompartmentEntity(rootCompartment.id, None, rootCompartment.name, None, 0, None)

    # Get those child compartments of root compartment - Level 1
    topCompartmentPayloadList = client.list_compartments(compartment_id=defaultProfileTenancy,
                                                sort_by="NAME",sort_order="ASC",
                                            #access_level="ANY",
                                            compartment_id_in_subtree=False).data
    topCompartmentList = []
    for child in topCompartmentPayloadList:
        compartmentEntity = CompartmentEntity(child.id, rootCompartment.id, child.name, rootCompartment.name, 1, None)
        topCompartmentList.append(compartmentEntity)

    # Get sub-compartments - Level 2
    subCompartmentList = []
    for parentCompartment in topCompartmentList:
        #print("List sub-compartments of compartment: ", parentCompartment["name"])
        payload = client.list_compartments(compartment_id=parentCompartment.id)
        childList = payload.data
        
        # childList is not empty
        for child in childList:
            compartmentEntity = CompartmentEntity(child.id, parentCompartment.id, child.name, parentCompartment.name, 2, parentCompartment.fullname)
            subCompartmentList.append(compartmentEntity)

    # Get child compartments of sub-compartment - Level 3
    childSubCompartmentList = []
    for parentCompartment in subCompartmentList:
        #print("List sub-compartments of compartment: ", parentCompartment["name"])
        payload = client.list_compartments(compartment_id=parentCompartment.id)
        childList = payload.data
        # childList is not empty
        for child in childList:
            compartmentEntity = CompartmentEntity(child.id, parentCompartment.id, child.name, parentCompartment.name, 3, parentCompartment.fullname)
            childSubCompartmentList.append(compartmentEntity)

    
    # Get grand child compartments of child-sub-compartment - Level 4
    grandChildSubCompartmentList = []
    for parentCompartment in childSubCompartmentList:
        #print("List sub-compartments of compartment: ", parentCompartment["name"])
        payload = client.list_compartments(compartment_id=parentCompartment.id)
        childList = payload.data
        # childList is not empty
        for child in childList:
            compartmentEntity = CompartmentEntity(child.id, parentCompartment.id, child.name, parentCompartment.name, 4, parentCompartment.fullname)
            grandChildSubCompartmentList.append(compartmentEntity)

    # Get grand child compartments of grand-child-sub-compartment - Level 5
    fifthGrandChildSubCompartmentList = []
    for parentCompartment in grandChildSubCompartmentList:
        #print("List sub-compartments of compartment: ", parentCompartment["name"])
        payload = client.list_compartments(compartment_id=parentCompartment.id)
        childList = payload.data
        # childList is not empty
        for child in childList:
            compartmentEntity = CompartmentEntity(child.id, parentCompartment.id, child.name, parentCompartment.name, 5, parentCompartment.fullname)
            fifthGrandChildSubCompartmentList.append(compartmentEntity)

    # Append the root compartment to list
    totalCompartments.append(rootCompartmentEntity)
    totalCompartments.extend(topCompartmentList)
    totalCompartments.extend(subCompartmentList)
    totalCompartments.extend(childSubCompartmentList)
    totalCompartments.extend(grandChildSubCompartmentList)
    totalCompartments.extend(fifthGrandChildSubCompartmentList)

    totalCompartments = totalCompartments[:30]

    for compartment in totalCompartments:
        #print(compartment.name)
        instance_count_group_by_compartment[compartment.name] = 0
        loadLalance_count_group_by_compartment[compartment.name] = 0
        mysql_count_group_by_compartment[compartment.name] = 0
        waf_count_group_by_compartment[compartment.name] = 0
        dbcs_count_group_by_compartment[compartment.name] = 0
        storage_count_group_by_compartment[compartment.name] = 0

    print("The number of compartments: ", len(totalCompartments))
    # print(totalCompartments)
    print("#***********************************************")
        
    # Begin to create a clients config list<dict>
    
    for subscribedRegion in subscribedRegionList:
        currentRegion = subscribedRegion.region_name
        clientConfig = generateClientConfig(defaultProfileUser, defaultProfileFingerprint, 
                                            defaultProfileKeyFile, defaultProfileTenancy, 
                                            currentRegion)
        clientConfigList.append(clientConfig)
    '''
    currentRegion = "ap-singapore-1"
    clientConfig = generateClientConfig(defaultProfileUser, defaultProfileFingerprint, 
                                        defaultProfileKeyFile, defaultProfileTenancy, 
                                        currentRegion)
    clientConfigList.append(clientConfig)

    currentRegion = "ap-tokyo-1"
    clientConfig = generateClientConfig(defaultProfileUser, defaultProfileFingerprint, 
                                        defaultProfileKeyFile, defaultProfileTenancy, 
                                        currentRegion)
    clientConfigList.append(clientConfig)

    currentRegion = "us-ashburn-1"
    clientConfig = generateClientConfig(defaultProfileUser, defaultProfileFingerprint, 
                                        defaultProfileKeyFile, defaultProfileTenancy, 
                                        currentRegion)
    clientConfigList.append(clientConfig)
    '''
    
    ## End to create a clients config list<dict>


def collectLoadBalances(clientConfigList, chckCompartmentList):
    print("Start-To-Collect-LoadBalances...")

    # Loop every client config then 
    for clientConfig in clientConfigList:
        currentRegion = clientConfig["region"]
        oci.config.validate_config(clientConfig)

        print("I'm here >>> ", currentRegion)
        
        load_balancer_client = oci.load_balancer.LoadBalancerClient(clientConfig)

        for compartment in totalCompartments:
            currentCompartmentName = compartment.name
            currentCompartmentId = compartment.id
            
            # ignore those compartments not in the checked list
            if isIgnoreCompartment(chckCompartmentList, compartment):
                continue
            
            list_load_balancers_response = load_balancer_client.list_load_balancers(
                compartment_id=currentCompartmentId)
            
            loadbalanceList = list_load_balancers_response.data
            for loadBalance in loadbalanceList:
                displayName = loadBalance.display_name
                state = loadBalance.lifecycle_state
                shape = loadBalance.shape_name

                metadata = {"isPrivate": loadBalance.is_private}
                data = ["LoadBalance", displayName, state, shape, currentCompartmentName, currentRegion, str(metadata)]
                all_resource_list.append(data)

            if(len(loadbalanceList) > 0):
                if(loadLalance_count_group_by_compartment.get(currentCompartmentName) == 0):
                    loadLalance_count_group_by_compartment[currentCompartmentName] = len(loadbalanceList)
                else:
                    loadLalance_count_group_by_compartment[currentCompartmentName] = loadLalance_count_group_by_compartment.get(currentCompartmentName) + len(loadbalanceList)
    print("End-To-Collect-LoadBalances!")

def collectComputeInstances(clientConfigList, chckCompartmentList):
    print("Start-To-Collect-Compute-Instances...")

    # Loop every client config then 
    for clientConfig in clientConfigList:
        currentRegion = clientConfig["region"]
        oci.config.validate_config(clientConfig)

        print("I'm here >>> ", currentRegion)
        
        #config = oci.config.from_file(profile_name=defaultProfile)
        computeClient = oci.core.ComputeClient(clientConfig)
        blockStorageClient = oci.core.BlockstorageClient(clientConfig)
        networkClient = oci.core.VirtualNetworkClient(clientConfig)

        for compartment in totalCompartments:
            currentCompartmentName = compartment.name
            currentCompartmentId = compartment.id
            
            # ignore those compartments not in the checked list
            if isIgnoreCompartment(chckCompartmentList, compartment):
                continue
            
            instancesResponse = computeClient.list_instances(currentCompartmentId)
            instanceList = instancesResponse.data
            #print("Region: {}, Compartment: {}, Number of Instances: {}".format(currentRegion, currentCompartmentName, len(instanceList)))
            for instance in instanceList:
                #print(instance)
                instanceId = instance.id
                instanceAd = instance.availability_domain
                '''
                bootVolumeDesc = ""
                bootVolumes = computeClient.list_boot_volume_attachments(availability_domain=instanceAd,compartment_id=currentCompartmentId,instance_id=instanceId).data
                if bootVolumes:
                    bvId = bootVolumes[0].boot_volume_id
                    try:
                        bootVolume = blockStorageClient.get_boot_volume(boot_volume_id=bvId).data
                        bootVolumeSize = bootVolume.size_in_gbs 
                        bootVolumeVPUs = bootVolume.vpus_per_gb
                        bootVolumeDesc = "{}(gb)-{}(pu)".format(bootVolumeSize, bootVolumeVPUs)
                    except:
                        print(">>>BootVolume Error:", bootVolumes)

                blockVolumeList = []
                blockVolumes = computeClient.list_volume_attachments(availability_domain=instanceAd,compartment_id=currentCompartmentId,instance_id=instanceId).data
                if blockVolumes:
                    for blv in blockVolumes:
                        #print(blv)
                        volumeId = blv.volume_id
                        try:
                            blockVolume = None
                            if volumeId.count('bootvolume') > 0:
                                blockVolume = blockStorageClient.get_boot_volume(boot_volume_id=volumeId).data
                            else:
                                blockVolume = blockStorageClient.get_volume(volume_id=volumeId).data

                            if blockVolume:
                                blockVolumeSize = blockVolume.size_in_gbs 
                                blockVolumeVPUs = blockVolume.vpus_per_gb
                                blockVolumeDesc = "{}(gb)-{}(pu)".format(blockVolumeSize, blockVolumeVPUs)
                                blockVolumeList.append(blockVolumeDesc)
                        except:
                            print(">>>Error in BlockVolume [ %s ]" % volumeId)
                
                blockVolumes = ""
                if blockVolumeList:
                    if len(blockVolumeList) == 1:
                        blockVolumes = blockVolumeList[0]
                    else:
                        blockVolumes = " & ".join(blockVolumeList)
                '''
                state = instance.lifecycle_state
                displayName = instance.display_name
                shapeConfig = instance.shape_config
                ocpus = shapeConfig.ocpus
                ram = shapeConfig.memory_in_gbs
                shape = instance.shape

                metadata = {"ocpu": ocpus, "memory": ram}
                data = ["virtualMachine", displayName, state, shape, currentCompartmentName, currentRegion, str(metadata)]
                all_resource_list.append(data)
                
                instance_count_group_by_compartment[currentCompartmentName] = instance_count_group_by_compartment.get(currentCompartmentName) + 1
    
    print("End-To-Collect-Compute-Instances!")

def collectAdbInstances(clientConfigList, chckCompartmentList, suffix):
    print("Start-To-Collect-Autonomouse-Databases...")
    # Define a csv file to store instances list
    prefix = "./adb-instances"
    filePath = "{}_{}.csv".format(prefix, suffix)
    instancesCsvFile = open(filePath, "a+")

    instancesCsvFile.write(f"Name, Status, OCPUs, Storage, Workload, Version, Compartment, Region\n")
    # Loop every client config then 
    for clientConfig in clientConfigList:
        currentRegion = clientConfig["region"]
        oci.config.validate_config(clientConfig)

        print("I'm here >>> ", currentRegion)
        
        #config = oci.config.from_file(profile_name=defaultProfile)
        databaseClient = oci.database.DatabaseClient(clientConfig)
        
        for compartment in totalCompartments:
            currentCompartmentName = compartment.fullname
            currentCompartmentId = compartment.id
            
            # ignore those compartments not in the checked list
            if isIgnoreCompartment(chckCompartmentList, compartment):
                continue
            
            
            # Query all Autonomous Databases
            adbResponse = databaseClient.list_autonomous_databases(compartment_id=currentCompartmentId)
            adbList = adbResponse.data
            
            # print("Region: {}, Compartment: {}, Number of Autonomous DB: {}".format(currentRegion, currentCompartmentName, len(adbList)))
            
            for instance in adbList:
                # print(instance)
                state = instance.lifecycle_state
                displayName = instance.display_name
                ocpus = instance.cpu_core_count
                storage = instance.data_storage_size_in_gbs
                workload = instance.db_workload
                version = instance.db_version
                # print(displayName, int(ocpus), int(storage), state)
                instancesCsvFile.write(f"{displayName},{state},{ocpus},{storage},{workload},{version},{currentCompartmentName},{currentRegion}\n")
                instancesCsvFile.flush()
            
    # close csv file
    instancesCsvFile.close()
    print("End-To-Collect-Autonomouse-Databases!")

def collectDBCSInstances(clientConfigList, chckCompartmentList):
    print("Start-To-Collect-DBCS-Instances...")
    # Define a csv file to store instances list
    # Loop every client config then 
    for clientConfig in clientConfigList:
        currentRegion = clientConfig["region"]
        oci.config.validate_config(clientConfig)

        print("I'm here >>> ", currentRegion)
        
        #config = oci.config.from_file(profile_name=defaultProfile)
        databaseClient = oci.database.DatabaseClient(clientConfig)
        
        for compartment in totalCompartments:
            currentCompartmentName = compartment.name
            currentCompartmentId = compartment.id
            
            # ignore those compartments not in the checked list
            if isIgnoreCompartment(chckCompartmentList, compartment):
                continue
            
            # Query all Managed Databases
            databasesResponse = databaseClient.list_db_systems(compartment_id=currentCompartmentId)
            dbInstanceList = databasesResponse.data
            #print(dbInstanceList)
            
            # print("Region: {}, Compartment: {}, Number of DBCS: {}".format(currentRegion, currentCompartmentName, len(dbInstanceList)))
            
            for instance in dbInstanceList:
                displayName = instance.display_name
                state = instance.lifecycle_state
                ocpus = instance.cpu_core_count
                storage = instance.data_storage_size_in_gbs
                shape = instance.shape
                edition = instance.database_edition
                version = instance.version

                metadata = {"version": version, "ocpus": ocpus}
                data = ["DBCS", displayName, state, shape, currentCompartmentName, currentRegion, str(metadata)]
                all_resource_list.append(data)
                
                dbcs_count_group_by_compartment[currentCompartmentName] = dbcs_count_group_by_compartment.get(currentCompartmentName) + 1
            
    print("End-To-Collect-DBCS-Instances!")

def collectMySQLInstances(clientConfigList, chckCompartmentList):
    print("Start-To-Collect-MySQL-Instances...")
    # Loop every client config then 
    for clientConfig in clientConfigList:
        currentRegion = clientConfig["region"]
        oci.config.validate_config(clientConfig)

        print("I'm here >>> ", currentRegion)
        
        #config = oci.config.from_file(profile_name=defaultProfile)
        mysqlClient = oci.mysql.DbSystemClient(clientConfig)
        mysqlIaaSClient = oci.mysql.MysqlaasClient(clientConfig)
        
        for compartment in totalCompartments:
            currentCompartmentName = compartment.name
            currentCompartmentId = compartment.id
            
            # ignore those compartments not in the checked list
            if isIgnoreCompartment(chckCompartmentList, compartment):
                continue
            
            # Query all Managed Databases
            mysqlResponse = mysqlClient.list_db_systems(compartment_id=currentCompartmentId,lifecycle_state='ACTIVE')
            mysqlList = mysqlResponse.data
                        
            for mysql in mysqlList:
                #print(mysql)
                displayName = mysql.display_name
                state = mysql.lifecycle_state
                version = mysql.mysql_version
                isHA = mysql.is_highly_available
                
                metadata = {"version": version, "isHA": isHA}
                data = ["MySQL", displayName, state, None, currentCompartmentName, currentRegion, str(metadata)]
                all_resource_list.append(data)
            
                mysql_count_group_by_compartment[currentCompartmentName] = mysql_count_group_by_compartment.get(currentCompartmentName) + 1
            
    print("End-To-Collect-MySQL-Instances!")

def collectWAFs(clientConfigList, chckCompartmentList):
    print("Start-To-Collect-WAF...")
    # Loop every client config then 
    for clientConfig in clientConfigList:
        currentRegion = clientConfig["region"]
        oci.config.validate_config(clientConfig)

        print("I'm here >>> ", currentRegion)
        
        #config = oci.config.from_file(profile_name=defaultProfile)
        waf_client = oci.waf.WafClient(clientConfig)
        
        for compartment in totalCompartments:
            currentCompartmentName = compartment.name
            currentCompartmentId = compartment.id
            
            # ignore those compartments not in the checked list
            if isIgnoreCompartment(chckCompartmentList, compartment):
                continue
            
            # Query all Managed Databases
            list_web_app_firewalls_response = waf_client.list_web_app_firewalls(
                compartment_id=currentCompartmentId)
            WAF_List = list_web_app_firewalls_response.data.items
                        
            for web_app_firewall in WAF_List:
                displayName = web_app_firewall.display_name
                state = web_app_firewall.lifecycle_state
                backend_type = web_app_firewall.backend_type
                
                metadata = {"backend_type": backend_type}
                data = ["WAF", displayName, state, None, currentCompartmentName, currentRegion, str(metadata)]
                all_resource_list.append(data)

                waf_count_group_by_compartment[currentCompartmentName] = waf_count_group_by_compartment.get(currentCompartmentName) + 1
            
    print("End-To-Collect-WAF-Instances!")

def collectStorages(clientConfigList, chckCompartmentList):
    print("Start-To-Collect-storages...")
    # Loop every client config then 
    for clientConfig in clientConfigList:
        currentRegion = clientConfig["region"]
        oci.config.validate_config(clientConfig)

        print("I'm here >>> ", currentRegion)
        
        #config = oci.config.from_file(profile_name=defaultProfile)
        object_storage_client = oci.object_storage.ObjectStorageClient(clientConfig)
        
        for compartment in totalCompartments:
            currentCompartmentName = compartment.name
            currentCompartmentId = compartment.id
            
            # ignore those compartments not in the checked list
            if isIgnoreCompartment(chckCompartmentList, compartment):
                continue
            
            # Query all Managed Databases
            list_buckets_response = object_storage_client.list_buckets(
                   compartment_id=currentCompartmentId,
                   namespace_name="sehubjapacprod")
            
            # Get the data from response
            bucketList = list_buckets_response.data
            for bucket in bucketList:
                displayName = bucket.name
                state = None
                namespace = bucket.namespace
                metadata = {"namespace": namespace}
                data = ["ObjectStorage", displayName, state, None, currentCompartmentName, currentRegion, str(metadata)]
                all_resource_list.append(data)
            
                storage_count_group_by_compartment[currentCompartmentName] = storage_count_group_by_compartment.get(currentCompartmentName) + 1
            
    print("End-To-Collect-storage!")

def write_excel(BYD_monthly_report):
    global instance_count_group_by_compartment
    global loadLalance_count_group_by_compartment
    global mysql_count_group_by_compartment
    global waf_count_group_by_compartment
    global dbcs_count_group_by_compartment
    global storage_count_group_by_compartment


    wb = load_workbook(BYD_monthly_report)

    resource_detail = wb.create_sheet("resource_detail")  #创建一个 sheet 名为 资源数量详情
    resource_detail.title = u"资源数量详情"  # 设置 sheet 标题

    resource_detail_heading = ["Resource_Type", "Resource_Name", "status", "shape", "Compartment", "Region", "metadata"]
    resource_detail.append(resource_detail_heading)

    resource_count = wb.create_sheet("resource_count")  #创建一个 sheet 名为 sheet
    resource_count.title = u"资源数量分类"  # 设置 sheet 标题
    
    for resource in all_resource_list:
        resource_detail.append(resource)

    compartments = [key for key in instance_count_group_by_compartment]
    instances_count = [value for value in instance_count_group_by_compartment.values()]
    resource_count.append([u"各区间虚拟机实例分布统计"])
    resource_count.append(compartments)
    resource_count.append(instances_count)
    resource_count.append([])


    compartments = [key for key in loadLalance_count_group_by_compartment]
    loadLalance_count = [value for value in loadLalance_count_group_by_compartment.values()]
    resource_count.append([u"各区间LoadBalance分布统计"])
    resource_count.append(compartments)
    resource_count.append(loadLalance_count)
    resource_count.append([])

    compartments = [key for key in mysql_count_group_by_compartment]
    mysql_count = [value for value in mysql_count_group_by_compartment.values()]
    resource_count.append([u"各区间MySQL服务分布统计"])
    resource_count.append(compartments)
    resource_count.append(mysql_count)
    resource_count.append([])

    compartments = [key for key in waf_count_group_by_compartment]
    waf_count = [value for value in waf_count_group_by_compartment.values()]
    resource_count.append([u"各区间WAF服务分布统计"])
    resource_count.append(compartments)
    resource_count.append(waf_count)
    resource_count.append([])

    compartments = [key for key in dbcs_count_group_by_compartment]
    dbcs_count = [value for value in dbcs_count_group_by_compartment.values()]
    resource_count.append([u"各区间DBCS服务分布统计"])
    resource_count.append(compartments)
    resource_count.append(dbcs_count)
    resource_count.append([])

    compartments = [key for key in storage_count_group_by_compartment]
    storage_count = [value for value in storage_count_group_by_compartment.values()]
    resource_count.append([u"各区间存储服务分布统计"])
    resource_count.append(compartments)
    resource_count.append(storage_count)
    resource_count.append([])

    # close csv file
    for i in [1,5,9,13,17,21]:
        range_string = "A1:M1"
        range_string = "A{}:M{}".format(i, i)
        resource_count.merge_cells(range_string=range_string)
    wb.save(BYD_monthly_report)
    

def collect_oci_resources(BYD_monthly_report):    
    # Only check those compartments you needs
    # [ 'SPECIALLIST2' ]
    chckCompartmentList = []
    initClientContext()

    # collect all instance resources group by compartment
    collectComputeInstances(clientConfigList, chckCompartmentList)

    # collect all storage resources group by compartment
    collectStorages(clientConfigList, chckCompartmentList)

    # collect all loadBalance resources group by compartment
    collectLoadBalances(clientConfigList, chckCompartmentList)

    # collect all waf resources group by compartment
    collectWAFs(clientConfigList, chckCompartmentList)

    # collect all DBCS resources group by compartment
    collectDBCSInstances(clientConfigList, chckCompartmentList)

    # collect all MySQL resources group by compartment
    collectMySQLInstances(clientConfigList, chckCompartmentList)

    # write all resources above to excel file
    write_excel(BYD_monthly_report=BYD_monthly_report)
