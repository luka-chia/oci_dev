# coding=utf-8
# This is an automatically generated code sample.
# To make this code sample work in your Oracle Cloud tenancy,
# please replace the values for any parameters whose current values do not fit
# your use case (such as resource IDs, strings containing ‘EXAMPLE’ or ‘unique_id’, and
# boolean, number, and enum parameters with values not fitting your use case).

import oci
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, colors, Alignment

def get_cost_report(tenant_id, BYD_monthly_report):
    # Initialize service client with default config file
    config = oci.config.from_file()
    usage_api_client = oci.usage_api.UsageapiClient(config)

    first_day_last_month = datetime.date(datetime.date.today().year,datetime.date.today().month-1,1)
    last_day_last_month = datetime.date(datetime.date.today().year,datetime.date.today().month,1)-datetime.timedelta(1)


    # Send the request to service, some parameters are not required, see API
    # doc for more info
    request_summarized_usages_response = usage_api_client.request_summarized_usages(
        request_summarized_usages_details=oci.usage_api.models.RequestSummarizedUsagesDetails(
            tenant_id=tenant_id,
            time_usage_started=datetime.datetime.strptime(str(first_day_last_month), "%Y-%m-%d"),
            time_usage_ended=datetime.datetime.strptime(str(datetime.date.today()), "%Y-%m-%d"),
            granularity="MONTHLY",
            group_by=["compartmentName","compartmentDepth"],
            query_type="COST",
            compartment_depth=2))

    # Get the data from response
    items = request_summarized_usages_response.data.items
    compartment_costs = dict()
    for item in items:
        if("INR"==item.currency):
            month = item.time_usage_started.strftime('%Y-%#m')
            cost = round(item.computed_amount,2)
            compartment_name = item.compartment_name
            
            if(cost is None):
                cost = 0
            if(compartment_name is None or compartment_name == " "):
                compartment_name = "no_compartment"

            if(compartment_name in compartment_costs):
                compartment_costs.get(compartment_name)[month] = cost
            else:
                month_cost = {month:cost}
                compartment_costs[compartment_name] = month_cost
    # print(compartment_costs)

    write_data(compartment_costs=compartment_costs, BYD_monthly_report=BYD_monthly_report)

def get_this_month():
	# 获取当前日期
    now_time = datetime.datetime.now()
    # 返回本月份
    return now_time.strftime('%Y-%#m')

def get_last_month():
	# 获取当前日期
    now_time = datetime.datetime.now()
    # 获取本月的第一天
    end_day_in_mouth = now_time.replace(day=1)
    # 获取上月的最后一天
    last_mouth = end_day_in_mouth - datetime.timedelta(days=1)
    # 返回上月的月份
    return last_mouth.strftime('%Y-%#m')

def parse_data(compartment_costs):
    compartment_names = [key for key in compartment_costs]

    last_month_cost = list()
    this_month_cost = list()

    for compartment_name in compartment_names:
        cost = compartment_costs.get(compartment_name)
        last_month_cost.append(cost.get(get_last_month(), 0))
        this_month_cost.append(cost.get(get_this_month(), 0))

    return compartment_names, last_month_cost, this_month_cost


def write_data(compartment_costs, BYD_monthly_report):
    compartment_names, last_month_cost, this_month_cost = parse_data(compartment_costs=compartment_costs)
    last_month = get_last_month()
    this_month = get_this_month()

    wb = load_workbook(BYD_monthly_report)

    ws = wb.create_sheet("cost_management")  #创建一个 sheet 名为 sheet
    ws.title = u"成本管理"  # 设置 sheet 标题

    compartment_names.insert(0, u"区间名及月份")
    last_month_cost.insert(0,last_month)
    this_month_cost.insert(0,this_month)
    rows = [ compartment_names, this_month_cost, last_month_cost ]

    for row in rows:
        ws.append(row)

    align = Alignment(horizontal='center', vertical='center',wrapText=False)
    # 两层循环遍历所有有数据的单元格
    for i in range(1, ws.max_row + 1):
        for j in range(1, ws.max_column + 1):
            ws.cell(i, j).alignment = align

    wb.save(BYD_monthly_report)
