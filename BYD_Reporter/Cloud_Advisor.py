# This is an automatically generated code sample.
# To make this code sample work in your Oracle Cloud tenancy,
# please replace the values for any parameters whose current values do not fit
# your use case (such as resource IDs, strings containing ‘EXAMPLE’ or ‘unique_id’, and
# boolean, number, and enum parameters with values not fitting your use case).

import oci
from openpyxl import load_workbook
from openpyxl.styles import Font, colors, Alignment

def get_advisor_report(compartment_id, BYD_monthly_report):
    # Initialize service client with default config file
    config = oci.config.from_file()
    optimizer_client = oci.optimizer.OptimizerClient(config)


    list_categories_response = optimizer_client.list_categories(
        compartment_id=compartment_id,
        compartment_id_in_subtree=True,
        # include_organization=True,
        opc_request_id="MSSQUNK7PGJX2EVRVUZM<unique_ID>")

    high_availability_advisor = dict()
    performance_advisor = dict()
    cost_advisor = dict()
    estimated_cost_saving = None
    # Get the data from response
    items = list_categories_response.data.items
    for item in items:
        name = item.name
        recommendation_data = {}
        for recommendation in item.recommendation_counts:
            recommendation_data[recommendation.importance] = recommendation.count
        
        resource_data = {}
        for resource in item.resource_counts:
            resource_data[resource.status] = resource.count

        if("high-availability" in name):
            # get high availability data
            high_availability_advisor["recommendation"] = recommendation_data
            high_availability_advisor["resource"] = resource_data
        elif("performance" in name):
            # get performance data
            performance_advisor["recommendation"] = recommendation_data
            performance_advisor["resource"] = resource_data
        elif("cost-management" in name):
            # get cost data
            cost_advisor["recommendation"] = recommendation_data
            cost_advisor["resource"] = resource_data
            estimated_cost_saving = item.estimated_cost_saving

    write_data(high_availability_advisor=high_availability_advisor, performance_advisor=performance_advisor,
                 cost_advisor=cost_advisor, estimated_cost_saving=estimated_cost_saving,
                 BYD_monthly_report=BYD_monthly_report)

def autopct_format(values):
    def my_format(pct):
        total = sum(values)
        val = int(round(pct*total/100.0))
        return '{v:d}'.format(v=val)
    return my_format

def transit_level(recommendation_level):
    en_cn_map = {
        "CRITICAL": u"严重",
        "HIGH":u"高",
        "MODERATE":u"中",
        "LOW":u"低",
        "MINOR":u"次要"
    }
    return en_cn_map.get(recommendation_level)

def write_data(high_availability_advisor, performance_advisor, cost_advisor, estimated_cost_saving, BYD_monthly_report):

    rows=list()
    rows.append([u"预计节省成本"])
    rows.append([estimated_cost_saving])
    rows.append([])

    high_availability_advisor_level = [transit_level(key) for key in high_availability_advisor.get("recommendation")]
    high_availability_advisor_count = [value for value in high_availability_advisor.get("recommendation").values()]
    rows.append([u"高可用建议"])
    rows.append(high_availability_advisor_level)
    rows.append(high_availability_advisor_count)
    rows.append([])

    performance_advisor_level = [transit_level(key) for key in performance_advisor.get("recommendation")]
    performance_advisor_count = [value for value in performance_advisor.get("recommendation").values()]
    rows.append([u"性能建议"])
    rows.append(performance_advisor_level)
    rows.append(performance_advisor_count)
    rows.append([])

    cost_advisor_level = [transit_level(key) for key in cost_advisor.get("recommendation")]
    cost_advisor_count = [value for value in cost_advisor.get("recommendation").values()]
    rows.append([u"成本建议"])
    rows.append(cost_advisor_level)
    rows.append(cost_advisor_count)

    wb = load_workbook(BYD_monthly_report)

    ws = wb.create_sheet("cloud_advisor")  #创建一个 sheet 名为 sheet
    ws.title = u"建议优化"  # 设置 sheet 标题

    #ws1.sheet_properties.tabColor = "#FF7F0E"  # 设置 sheet 标签背景色

    for row in rows:
        ws.append(row)

    ws.merge_cells(range_string="A4:E4")
    ws.merge_cells(range_string="A8:E8")
    ws.merge_cells(range_string="A12:E12")

    align = Alignment(horizontal='center', vertical='center',wrapText=False)
    # 两层循环遍历所有有数据的单元格
    for i in range(1, ws.max_row + 1):
        for j in range(1, ws.max_column + 1):
            ws.cell(i, j).alignment = align

    wb.save(BYD_monthly_report)
