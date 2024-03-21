# This is an automatically generated code sample.
# To make this code sample work in your Oracle Cloud tenancy,
# please replace the values for any parameters whose current values do not fit
# your use case (such as resource IDs, strings containing ‘EXAMPLE’ or ‘unique_id’, and
# boolean, number, and enum parameters with values not fitting your use case).

import oci
import datetime
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import Font, colors, Alignment

# Create a default config using DEFAULT profile in default location
# Refer to
# https://docs.cloud.oracle.com/en-us/iaas/Content/API/Concepts/sdkconfig.htm#SDK_and_CLI_Configuration_File
# for more info

def get_cloud_guard_report(compartment_id, BYD_monthly_report):
    # Initialize service client with default config file
    config = oci.config.from_file()
    cloud_guard_client = oci.cloud_guard.CloudGuardClient(config)


    # Send the request to get the risk_scores
    request_risk_scores_response = cloud_guard_client.request_risk_scores(
        compartment_id=compartment_id,
        limit=49)
    risk_score = request_risk_scores_response.data.items[0].risk_score
    # print("Currently the tenancy risk_scores = %s" % risk_score)


    # Send the request to get the security_scores
    request_security_scores_response = cloud_guard_client.request_security_scores(
        compartment_id=compartment_id,
        limit=235)
    security_score = request_security_scores_response.data.items[0].security_score
    #print("Currently the tenancy security_scores = %s, and security_rating = %s"
    #    %(security_score, request_security_scores_response.data.items[0].security_rating))

    # Send the request to get problems
    request_summarized_problems_response = cloud_guard_client.request_summarized_problems(
        list_dimensions=["RISK_LEVEL"],
        compartment_id=compartment_id,
        limit=6910,
        compartment_id_in_subtree=True,
        access_level="ACCESSIBLE")

    problem_items = request_summarized_problems_response.data.items
    problem_data = {}
    for item in problem_items:
        problem_data[item.dimensions_map.get("RISK_LEVEL")] = item.count
    # print(problem_data)

    # Send the request to get the SecurityScoreSummarizedTrend
    end = datetime.datetime.now()
    begin = end-datetime.timedelta(days=30)
    request_security_score_summarized_trend_response = cloud_guard_client.request_security_score_summarized_trend(
        compartment_id=compartment_id,
        time_score_computed_greater_than_or_equal_to=begin,
        time_score_computed_less_than_or_equal_to=end)

    trend_items = request_security_score_summarized_trend_response.data.items
    security_score_trend = {}
    for item in trend_items:
        time_stamp = datetime.datetime.fromtimestamp(item.start_timestamp/1000)
        security_score_trend[time_stamp.strftime("%Y-%m-%d")] = item.security_score
        #print("the security_score = %s and security_rating = %s in %s"
        #    %(item.security_score, item.security_rating, datetime.datetime.fromtimestamp(item.start_timestamp/1000)))
    # print(security_score_trend)

    write_data(risk_score=risk_score, security_score=security_score,
                 problem_data=problem_data, security_score_trend=security_score_trend,
                 BYD_monthly_report=BYD_monthly_report)

def autopct_format(values):
    def my_format(pct):
        total = sum(values)
        val = int(round(pct*total/100.0))
        return '{v:d}'.format(v=val)
    return my_format

def transit_risk_level(risk_level):
    en_cn_map = {
        "CRITICAL": u"严重",
        "HIGH":u"高",
        "MEDIUM":u"中",
        "LOW":u"低",
        "MINOR":u"次要"
    }
    return en_cn_map.get(risk_level)

def write_data(risk_score, security_score, problem_data, security_score_trend, BYD_monthly_report):
    rows=list()
    risk_score_row = [u"风险等级评分", risk_score]
    rows.append(risk_score_row)
    rows.append([])

    security_score_row = [u"安全等级评分", security_score]
    rows.append(security_score_row)
    rows.append([])

    risk_level = [transit_risk_level(key) for key in problem_data]
    problem_count = [value for value in problem_data.values()]
    rows.append([u"不同安全等级问题分类"])
    rows.append(risk_level)
    rows.append(problem_count)
    rows.append([])

    # step4 draw link picture about security score trend
    time_stamp = [key for key in security_score_trend]
    security_score = [value for value in security_score_trend.values()]
    rows.append([u"近30天安全评分趋势"])
    rows.append(time_stamp)
    rows.append(security_score)
    rows.append([])

    wb = load_workbook(BYD_monthly_report)

    ws = wb.create_sheet("cloud_guard")  #创建一个 sheet 名为 sheet
    ws.title = u"安全评估"  # 设置 sheet 标题
    

    for row in rows:
        ws.append(row)
    ws.merge_cells(range_string="A5:E5")
    ws.merge_cells(range_string="A9:E9")

    align = Alignment(horizontal='center', vertical='center',wrapText=False)
    # 两层循环遍历所有有数据的单元格
    for i in range(1, ws.max_row + 1):
        for j in range(1, ws.max_column + 1):
            ws.cell(i, j).alignment = align

    wb.save(BYD_monthly_report)
