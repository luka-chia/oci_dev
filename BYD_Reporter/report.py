import os
import Cloud_Guard
import Cost_Management
import Cloud_Advisor
import Resource_Utilization
from OCI_tools import CollectInstances
from openpyxl import Workbook
from openpyxl import load_workbook


compartment_id = "ocid1.tenancy.oc1..aaaaaaaaro7aox2fclu4urtpgsbacnrmjv46e7n4fw3sc2wbq24l7dzf3kba"
tenant_id="ocid1.tenancy.oc1..aaaaaaaaro7aox2fclu4urtpgsbacnrmjv46e7n4fw3sc2wbq24l7dzf3kba"

BYD_monthly_report = "/Users/luka/BYD_monthly_report.xlsx"

# 新建excel空文件
# wb = Workbook()

# 加载已有目标文件
wb = load_workbook(BYD_monthly_report)
# 要删除的Sheet表名称列表
sheet_delete_list = [u"成本管理", u"安全评估", u"建议优化", u"资源使用率", u"资源数量详情", u"资源数量分类"]

# 遍历并删除Sheet表
for sheet_name in sheet_delete_list:
    try:
        sheet_delete = wb[sheet_name]#wb.get_sheet_by_name(sheet_name)
        wb.remove(sheet_delete)
    except:
        continue

wb.save(BYD_monthly_report)

# get resource count
print("==== step 0: get resource count by compartment ====")
CollectInstances.collect_oci_resources(BYD_monthly_report=BYD_monthly_report)

# get cost report by compartment from Cost_Management service
print("==== step 1: get cost report by compartment from Cost_Management service ====")
Cost_Management.get_cost_report(tenant_id=tenant_id, BYD_monthly_report=BYD_monthly_report)

# get security report from Cloud_Guard service
print("==== step 2: get security report from Cloud_Guard service ====")
Cloud_Guard.get_cloud_guard_report(compartment_id=compartment_id, BYD_monthly_report=BYD_monthly_report)

# get advisor report from Cloud_Advisor service
print("==== step 3: get advisor report from Cloud_Advisor service ====")
Cloud_Advisor.get_advisor_report(compartment_id=compartment_id, BYD_monthly_report=BYD_monthly_report)

# get resource utilization report from operation insight service
print("==== step 4: get resource utilization report from operation insight service ====")
Resource_Utilization.get_resource_utilization_report(compartment_id=compartment_id, 
                                                     BYD_monthly_report=BYD_monthly_report)
