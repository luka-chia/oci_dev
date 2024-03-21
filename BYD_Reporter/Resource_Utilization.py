import oci
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, colors, Alignment

def get_resource_utilization_report(compartment_id, BYD_monthly_report):
    config = oci.config.from_file()
    config.update({"region": "ap-seoul-1"})


    # Initialize service client with default config file
    opsi_client = oci.opsi.OperationsInsightsClient(config)

    # get cpu statistics
    cpu_statistics = opsi_client.summarize_host_insight_resource_statistics(
        compartment_id=compartment_id,
        resource_metric="CPU",
        analysis_time_interval="P30D",
        limit=207,
        sort_order="DESC",
        sort_by="utilizationPercent",
        opc_request_id="B5EG7RA4QQS76JRGMROY<unique_ID>",
        compartment_id_in_subtree=True)

    # Get the data from response
    cpu_items = cpu_statistics.data.items
    topN_by_cpu = {}
    # print("=============================== cpu statistics ==============================")
    for item in cpu_items:
        resource_name = deal(item.host_details.host_display_name)
        utilization_percent = item.current_statistics.utilization_percent
        topN_by_cpu[resource_name] = utilization_percent
        #print("host[name=%s] cpu_usage_percent=%s"%(item.host_details.host_display_name,item.current_statistics.utilization_percent))
    # print(topN_by_cpu)

    # get mem statistics
    memory_statistics = opsi_client.summarize_host_insight_resource_statistics(
        compartment_id=compartment_id,
        resource_metric="LOGICAL_MEMORY",
        analysis_time_interval="P30D",
        limit=207,
        sort_order="DESC",
        sort_by="utilizationPercent",
        opc_request_id="B5EG7RA4QQS76JRGMROY<unique_ID>",
        compartment_id_in_subtree=True)

    # Get the data from response
    memory_items = memory_statistics.data.items
    topN_by_memory = {}
    # print("=============================== memory statistics ==============================")
    for item in memory_items:
        resource_name = deal(item.host_details.host_display_name)
        utilization_percent = item.current_statistics.utilization_percent
        topN_by_memory[resource_name] = utilization_percent
        #print("host[name=%s] memory_usage_percent=%s"%(item.host_details.host_display_name,item.current_statistics.utilization_percent))
    # print(topN_by_memory)

    # get storage statistics
    storage_statistics = opsi_client.summarize_host_insight_resource_statistics(
        compartment_id=compartment_id,
        resource_metric="STORAGE",
        analysis_time_interval="P30D",
        limit=207,
        sort_order="DESC",
        sort_by="utilizationPercent",
        opc_request_id="B5EG7RA4QQS76JRGMROY<unique_ID>",
        compartment_id_in_subtree=True)

    # Get the data from response
    storage_items = storage_statistics.data.items
    topN_by_storage = {}
    # print("=============================== storage statistics ==============================")
    for item in storage_items:
        resource_name = deal(item.host_details.host_display_name)
        utilization_percent = item.current_statistics.utilization_percent
        topN_by_storage[resource_name] = utilization_percent
        #print("host[name=%s] storage_usage_percent=%s"%(item.host_details.host_display_name,item.current_statistics.utilization_percent))
    # print(topN_by_storage)

    write_data(topN_by_cpu=topN_by_cpu, topN_by_memory=topN_by_memory, topN_by_storage=topN_by_storage, 
               BYD_monthly_report=BYD_monthly_report)

def deal(string):
    return string.split(".")[0]

def write_data(topN_by_cpu, topN_by_memory, topN_by_storage, BYD_monthly_report):

    rows=list()
    resource_name = [key for key in topN_by_cpu]
    cpu_percent = [round(value, 2) for value in topN_by_cpu.values()]

    rows.append([u"CPU资源使用率排行"])
    rows.append(resource_name[::-1])
    rows.append(cpu_percent[::-1])
    rows.append([])

    rows.append([u"内存资源使用率排行"])
    resource_name = [key for key in topN_by_memory]
    memory_percent = [round(value, 2) for value in topN_by_memory.values()]
    rows.append(resource_name[::-1])
    rows.append(memory_percent[::-1])
    rows.append([])

    rows.append([u"存储资源使用率排行"])
    resource_name = [key for key in topN_by_storage]
    storage_percent = [round(value, 2) for value in topN_by_storage.values()]
    rows.append(resource_name[::-1])
    rows.append(storage_percent[::-1])

    wb = load_workbook(BYD_monthly_report)

    ws = wb.create_sheet("resource_utilization")  #创建一个 sheet 名为 sheet
    ws.title = u"资源使用率"  # 设置 sheet 标题

    for row in rows:
        ws.append(row)
    
    ws.merge_cells(range_string="A1:E1")
    ws.merge_cells(range_string="A5:E5")
    ws.merge_cells(range_string="A9:E9")

    align = Alignment(horizontal='center', vertical='center',wrapText=False)
    # 两层循环遍历所有有数据的单元格
    for i in range(1, ws.max_row + 1):
        for j in range(1, ws.max_column + 1):
            ws.cell(i, j).alignment = align

    wb.save(BYD_monthly_report)
