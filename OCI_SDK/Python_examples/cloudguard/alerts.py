# This is an automatically generated code sample.
# To make this code sample work in your Oracle Cloud tenancy,
# please replace the values for any parameters whose current values do not fit
# your use case (such as resource IDs, strings containing ‘EXAMPLE’ or ‘unique_id’, and
# boolean, number, and enum parameters with values not fitting your use case).

import oci
import datetime
import json
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font, colors, Alignment

totalCompartments = list()
totalCompartments_name_id_map = dict()

def init_env(compartment_id):
    config = oci.config.from_file()
    client = oci.identity.IdentityClient(config)
    # Get the root compartment - Level 0
    rootCompartment = client.get_compartment(compartment_id=compartment_id).data
    totalCompartments_name_id_map[rootCompartment.id] = rootCompartment.name

    # Get those child compartments of root compartment - Level 1
    level_1_compartments = client.list_compartments(compartment_id=compartment_id,
                                                sort_by="NAME",sort_order="ASC",
                                            compartment_id_in_subtree=False).data
    level_1_compartment_ids = list()
    level_2_compartment_ids = list()
    level_3_compartment_ids = list()
    level_4_compartment_ids = list()
    level_5_compartment_ids = list()
    for child in level_1_compartments:
        totalCompartments_name_id_map[child.id] = child.name
        level_1_compartment_ids.append(child.id)

    # Get those child compartments of root compartment - Level 2
    for item_id in level_1_compartment_ids:
        level_2_compartments = client.list_compartments(compartment_id=item_id,
                                                    sort_by="NAME",sort_order="ASC",
                                                compartment_id_in_subtree=False).data
        for child in level_2_compartments:
            totalCompartments_name_id_map[child.id] = child.name
            level_2_compartment_ids.append(child.id)
    
    # Get those child compartments of root compartment - Level 3
    for item_id in level_2_compartment_ids:
        level_3_compartments = client.list_compartments(compartment_id=item_id,
                                                    sort_by="NAME",sort_order="ASC",
                                                compartment_id_in_subtree=False).data
        for child in level_3_compartments:
            totalCompartments_name_id_map[child.id] = child.name
            level_3_compartment_ids.append(child.id)
    
    # Get those child compartments of root compartment - Level 4
    for item_id in level_3_compartment_ids:
        level_4_compartments = client.list_compartments(compartment_id=item_id,
                                                    sort_by="NAME",sort_order="ASC",
                                                compartment_id_in_subtree=False).data
        for child in level_4_compartments:
            totalCompartments_name_id_map[child.id] = child.name
            level_4_compartment_ids.append(child.id)
    
    # Get those child compartments of root compartment - Level 5
    for item_id in level_4_compartment_ids:
        level_5_compartments = client.list_compartments(compartment_id=item_id,
                                                    sort_by="NAME",sort_order="ASC",
                                                compartment_id_in_subtree=False).data
        for child in level_5_compartments:
            totalCompartments_name_id_map[child.id] = child.name
            level_5_compartment_ids.append(child.id)
    
    #print("The number of compartments: ", len(totalCompartments_name_id_map))
    #print(totalCompartments_name_id_map)


def get_cloud_guard_report(compartment_id, BYD_CloudGuard_Problems_Report):
    # Initialize service client with default config file
    config = oci.config.from_file()
    cloud_guard_client = oci.cloud_guard.CloudGuardClient(config)

    list_problems_response = cloud_guard_client.list_problems(
        compartment_id=compartment_id,
        compartment_id_in_subtree=True,
        access_level="ACCESSIBLE",
        detector_rule_id_list=["SECURITY_LISTS_OPEN_PORTS"],
        limit=200)

    # Get the data from response
    problems = []
    for item in list_problems_response.data.items:
        problem_id = item.id
        get_problem_response = cloud_guard_client.get_problem(
            problem_id=problem_id,
            opc_request_id="OHVGHBBLWRPMBBGEQ8KV<unique_ID>")

        # Get the data from response
        problem = get_problem_response.data
        problem_detail = {
            "resource": problem.resource_name,
            "resource_id": problem.resource_id,
            "resource_type": problem.resource_type,
            "risk_level": transit_risk_level(problem.risk_level),
            "region": problem.region,
            "compartment_id": totalCompartments_name_id_map.get(problem.compartment_id,"luka"),
            "additional_details": json.dumps(problem.additional_details),
        }
        problems.append(problem_detail)
    
    print(problems)

    write_data(problems=problems,
               BYD_CloudGuard_Problems_Report=BYD_CloudGuard_Problems_Report)

def autopct_format(values):
    def my_format(pct):
        total = sum(values)
        val = int(round(pct*total/100.0))
        return '{v:d}'.format(v=val)
    return my_format

def transit_risk_level(risk_level):
    en_cn_map = {
        "CRITICAL": u"严重",
        "HIGH":u"高",
        "MEDIUM":u"中",
        "LOW":u"低",
        "MINOR":u"次要"
    }
    return en_cn_map.get(risk_level)

def write_data(problems, BYD_CloudGuard_Problems_Report):
    
    wb = load_workbook(BYD_CloudGuard_Problems_Report)

    default_sheet = wb.active
    default_sheet.title = "cloud_guard_problem"
    
    row = ["资源名称","资源ID","资源类型","风险级别","Region","区间","详细信息"]
    default_sheet.append(row)

    for problem in problems:
        row = [value for value in problem.values()]
        default_sheet.append(row)

    align = Alignment(horizontal='center', vertical='center',wrapText=False)
    # 两层循环遍历所有有数据的单元格
    for i in range(1, default_sheet.max_row + 1):
        for j in range(1, default_sheet.max_column + 1):
            default_sheet.cell(i, j).alignment = align

    wb.save(BYD_CloudGuard_Problems_Report)

BYD_CloudGuard_Problems_Report = "/Users/luka/BYD_CloudGuard_Problems.xlsx"
# 新建excel空文件
wb = Workbook()
wb.save(BYD_CloudGuard_Problems_Report)

# my sehubjapacprod
# compartment_id = "ocid1.tenancy.oc1..aaaaaaaaro7aox2fclu4urtpgsbacnrmjv46e7n4fw3sc2wbq24l7dzf3kba"
# byd
compartment_id = "ocid1.tenancy.oc1..aaaaaaaak5xfg2cv2jdtdj37sbknj2r47ma2t3gyfaljqnfp4sbgt2iv2r4q"

init_env(compartment_id=compartment_id)
get_cloud_guard_report(compartment_id=compartment_id, BYD_CloudGuard_Problems_Report=BYD_CloudGuard_Problems_Report)
